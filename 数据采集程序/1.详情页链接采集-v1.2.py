import time
import openpyxl
from selenium.webdriver.common.by import By
import pandas as pd
from tools import *
from selenium import webdriver
from selenium.webdriver.edge.service import Service
from selenium.webdriver.edge.options import Options
import requests
import random
import json
import os
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

links = []

# 获取数据文件夹下的所有文件
data_folder = '../数据/'
all_files = os.listdir(data_folder)

# 只处理前20个文件
files_to_process = all_files[:20]

for file_name in files_to_process:
    if file_name.endswith('.csv'):
        path = os.path.join(data_folder, file_name)
        df = pd.read_csv(path)

        # 将数据写入Excel文件
        excel_file = f'../xlsx/{file_name[:-4]}.xlsx'
        df.to_excel(excel_file, index=False)
        wb = openpyxl.load_workbook(excel_file)
        sheet = wb['Sheet1']
        # 获取总行数
        dim = sheet.dimensions.split(':')
        for i in range(2, int(dim[1][1:]) + 1):
            links.append(sheet[f'A{i}'].value)

# 在文件顶部添加cookie文件列表
BOSS_COOKIE_FILES = [
    '../cookie/bosscookies1.txt',
    '../cookie/bosscookies2.txt',
    '../cookie/bosscookies3.txt',
    'D:\\Project_Files\\Projects_Pycharm\爬虫\\boss数据爬取阶段\\cookie\\bosscookies4.txt',
    'D:\\Project_Files\\Projects_Pycharm\爬虫\\boss数据爬取阶段\\cookie\\bosscookies5.txt',
    'D:\\Project_Files\\Projects_Pycharm\爬虫\\boss数据爬取阶段\\cookie\\bosscookies6.txt',
]

def load_cookies(driver, cookie_file):
    """加载指定cookie文件"""
    try:
        with open(cookie_file, 'r', encoding='utf-8') as f:
            cookies = json.load(f)
        # 先访问与cookie相关的域名以设置cookie
        driver.get('https://www.zhipin.com')
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.TAG_NAME, 'body')))
        for cookie in cookies:
            # 处理过期时间格式
            if 'expiry' in cookie:
                cookie['expiry'] = int(cookie['expiry'])
            driver.add_cookie(cookie)
    except Exception as e:
        print(f"加载cookie失败: {str(e)}")

def get_driver(proxy=None):
    """创建Edge WebDriver实例并加载随机cookie"""
    service = Service(executable_path='D:/Software_work/Python312/msedgedriver.exe')  # 使用msedgedriver
    options = Options()
    if proxy:
        options.add_argument(f'--proxy-server={proxy}')
    driver = webdriver.Edge(service=service, options=options)
    
    # 随机选择cookie文件
    cookie_file = random.choice(BOSS_COOKIE_FILES)
    load_cookies(driver, cookie_file)
    
    return driver

def get_new_proxy():
    """通过API获取新的代理IP"""
    response = requests.get("http://www.xiongmaodaili.com/xiongmao-web/api/glip?secret=a5eba7e1d43d8630e713712a2199572a&orderNo=GL20250219082927NXfIvHlh&count=1&isTxt=0&proxyType=1&returnAccount=1")
    if response.status_code == 200:
        data = response.json()
        if data["code"] == "0":
            ip = data["obj"][0]["ip"]
            port = data["obj"][0]["port"]
            return f"{ip}:{port}"
    print("无法获取新的代理IP")
    return None

for i in range(len(links)):
    data = []
    proxy = get_new_proxy()
    driver = get_driver(proxy)
    driver.get(links[i])
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.TAG_NAME, 'body')))

    # 检测403错误
    if driver.title == "403 Forbidden":
        print("检测到403错误，检查特定元素")
        try:
            # 检查特定的XPATH元素
            element = WebDriverWait(driver, 5).until(
                EC.presence_of_element_located((By.XPATH, '//*[@id="code31"]/div[1]/p[2]/span[2]/a'))
            )
            if element:
                print("找到特定元素，点击该元素")
                element.click()
                WebDriverWait(driver, 5).until(EC.staleness_of(element))
        except:
            print("未找到特定元素，切换代理")
            driver.close()
            proxy = get_new_proxy()
            driver = get_driver(proxy)
            driver.get(links[i])
            WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.TAG_NAME, 'body')))

    # 检测登录提示
    try:
        login_prompt = WebDriverWait(driver, 5).until(
            EC.presence_of_element_located((By.XPATH, '/html/body/div/div[1]/div/div/div[1]/div[1]/p[2]/span[2]/a'))
        )
        if login_prompt:
            print("检测到登录提示，切换代理")
            driver.close()
            proxy = get_new_proxy()
            driver = get_driver(proxy)
            driver.get(links[i])
            WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.TAG_NAME, 'body')))
    except:
        pass

    jishuqi = 0
    while True:
        rawdata = driver.find_elements(by=By.XPATH, value="//a[@class='job-card-left']")
        jishuqi += 1
        time.sleep(0.4)
        if jishuqi == 10:
            break
        if rawdata != []:
            break

    try:
        kong = WebDriverWait(driver, 5).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "#wrap > div.page-job-wrapper > div.page-job-inner > div > div.job-list-wrapper > div.search-job-result.job-result-empty > div > div > p"))
        )
        if kong.text == '没有找到相关职位，打开 APP，查看全部职位库，优质职位随心聊。':
            print(path, f'第{i+1}个链接', '内容为空')
            driver.close()
            continue
    except:
        pass

    for j in rawdata:
        data.append(j.get_attribute('href'))

    if not os.path.exists(f'../results/{path[7:-4]}.xlsx'):
        # 如果文件不存在，创建文件
        wb = openpyxl.Workbook()
        ws = wb.active
        wb.save(f'../results/{path[7:-4]}.xlsx')

    for m in data:
        wb = openpyxl.load_workbook(f'../results/{path[7:-4]}.xlsx')
        # 获取指定的sheet
        ws = wb["Sheet"]
        max_row_num = ws.max_row
        ws._current_row = max_row_num
        ws.append([m])
        wb.save(f'../results/{path[7:-4]}.xlsx')
    if rawdata == []:
        print(path, f'第{i+1}个链接', '未完成')
    else:
        print(path, f'第{i+1}个链接', '已完成')

    driver.close()