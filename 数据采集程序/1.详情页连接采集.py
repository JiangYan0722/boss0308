import time
import openpyxl
from selenium.webdriver.common.by import By
import pandas as pd
from tools import *
from tqdm import tqdm, trange
import os
from selenium.webdriver.edge.options import Options
from selenium.webdriver.edge.service import Service
from selenium.webdriver.edge.webdriver import WebDriver
import random
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import requests
import logging

links = []

# 设置日志记录
logging.basicConfig(filename='../403/403_errors.log', level=logging.INFO, format='%(asctime)s - %(message)s')

def log_403_error(link):
    """记录403错误的链接"""
    logging.info(f"403 Error: {link}")

# 1. 自动扫描数据目录下所有CSV文件
data_directory = '../数据/'
csv_files = [f for f in os.listdir(data_directory) if f.endswith('.csv')]

def get_driver_with_proxy(proxy):
    """使用指定的代理创建Edge WebDriver实例"""
    options = Options()
    options.add_argument(f'--proxy-server={proxy}')
    # options.add_argument('--headless')  # 启用无头模式
    options.add_argument('--disable-gpu')  # 如果系统支持GPU加速，禁用它
    options.add_argument('--no-sandbox')  # 解决DevToolsActivePort文件不存在的错误
    service = Service(executable_path='D:/Software_work/Python312/msedgedriver.exe')  # 设置为实际路径
    driver = WebDriver(service=service, options=options)
    return driver

def get_new_proxy():
    """通过API获取新的代理IP"""
    response = requests.get("http://www.xiongmaodaili.com/xiongmao-web/api/glip?secret=a5eba7e1d43d8630e713712a2199572a&orderNo=GL20250219082927NXfIvHlh&count=1&isTxt=0&proxyType=1&returnAccount=1")
    if response.status_code == 200:
        data = response.json()
        if data["code"] == "0":
            ip = data["obj"][0]["ip"]
            port = data["obj"][0]["port"]
            return f"http://{ip}:{port}"
    print("无法获取新的代理IP")
    return None

def switch_proxy_and_retry(driver, link):
    """切换代理并重试"""
    driver.close()
    new_proxy = get_new_proxy()
    if new_proxy:
        print(f"切换到新代理: {new_proxy}")
        driver = get_driver_with_proxy(new_proxy)
        driver.get(link)
    else:
        print("无法获取新代理，跳过此链接")
    return driver

def send_email_notification():
    """发送邮箱验证码提示"""
    sender_email = "your_email@example.coom" 
    receiver_email = "1917811712@qq.com"
    password = "your_email_password"

    message = MIMEMultipart()
    message["From"] = sender_email
    message["To"] = receiver_email
    message["Subject"] = "验证码提示"

    body = "页面出现验证码图片，请手动验证。"
    message.attach(MIMEText(body, "plain"))

    try:
        server = smtplib.SMTP("smtp.example.com", 587)
        server.starttls()
        server.login(sender_email, password)
        server.sendmail(sender_email, receiver_email, message.as_string())
        server.close()
        print("邮件发送成功")
    except Exception as e:
        print(f"邮件发送失败: {e}")

def click_login_button(driver):
    """点击登录按钮"""
    try:
        login_button = driver.find_element(by=By.XPATH, value='//*[@id="code31"]/div[1]/p[2]/span[2]/a')
        login_button.click()
        print("点击登录按钮")
    except Exception as e:
        print(f"无法点击登录按钮: {e}")

for csv_file in csv_files:
    path = os.path.join(data_directory, csv_file)
    df = pd.read_csv(path)

    # 将每个文件中的链接添加到列表中
    for link in df['url']:  # 假设CSV文件中有一列名为'url'
        links.append(link)

    # 2. 准备写入将数据写入Excel文件
    excel_file = f'../xlsx/{csv_file[:-4]}.xlsx'
    df.to_excel(excel_file, index=False)
    wb = openpyxl.load_workbook(excel_file)
    sheet = wb['Sheet1']
    # 获取总行数
    dim = sheet.dimensions.split(':')
    for i in range(2, int(dim[1][1:]) + 1):
        links.append(sheet[f'A{i}'].value)

    for i in tqdm(range(len(links)), desc='...' ):
        data = []
        proxy = get_new_proxy()  # 动态获取代理
        if not proxy:
            print("无法获取代理，跳过此链接")
            continue
        driver = get_driver_with_proxy(proxy)
        driver.get(links[i])
        time.sleep(3)

        # 检测403错误
        if driver.title == "403 Forbidden":
            print("检测到403错误")
            log_403_error(links[i])
            click_login_button(driver)  # 点击登录按钮
            driver = switch_proxy_and_retry(driver, links[i])
            continue

        # 检测验证码图片
        try:
            captcha = driver.find_element(by=By.XPATH, value="//img[@alt='验证码']")
            if captcha:
                print("检测到验证码图片")
                send_email_notification()
                driver = switch_proxy_and_retry(driver, links[i])
                continue
        except:
            pass

        # 检测"您暂时无法继续访问~"文本
        try:
            block_message = driver.find_element(by=By.XPATH, value="//*[contains(text(), '您暂时无法继续访问~')]")
            if block_message:
                print("检测到访问限制提示")
                driver = switch_proxy_and_retry(driver, links[i])
                continue
        except:
            pass

        jishuqi = 0
        while True:
            rawdata = driver.find_elements(by=By.XPATH, value="//a[@class='job-card-left']")
            jishuqi += 1
            time.sleep(0.4)
            if jishuqi == 10:
                break
            if rawdata != []:
                break

        try:
            kong = driver.find_element(by=By.CSS_SELECTOR, value="#wrap > div.page-job-wrapper > div.page-job-inner > div > div.job-list-wrapper > div.search-job-result.job-result-empty > div > div > p")
            if kong.text == '没有找到相关职位，打开 APP，查看全部职位库，优质职位随心聊。':
                print(path, f'第{i+1}个链接', '内容为空')
                driver.close()
                continue
        except:
            pass

        for j in rawdata:
            data.append(j.get_attribute('href'))

        if not os.path.exists(f'../results/{csv_file[:-4]}.xlsx'):
            # 如果文件不存在，创建文件
            wb = openpyxl.Workbook()
            ws = wb.active
            wb.save(f'../results/{csv_file[:-4]}.xlsx')

        for m in data:
            wb = openpyxl.load_workbook(f'../results/{csv_file[:-4]}.xlsx')
            # 获取指定的sheet
            ws = wb["Sheet"]
            max_row_num = ws.max_row
            ws._current_row = max_row_num
            ws.append([m])
            wb.save(f'../results/{csv_file[:-4]}.xlsx')
        if rawdata == []:
            print(path, f'第{i+1}个链接', '未完成')
        else:
            print(path, f'第{i+1}个链接', '已完成')

        driver.close()




# path = 'https://www.zhipin.com/web/geek/job?query=%E6%AC%A3%E6%97%BA%E8%BE%BE%E7